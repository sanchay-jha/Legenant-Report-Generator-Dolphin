import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

# ----------------------------
# APP CONFIG
# ----------------------------
st.set_page_config(page_title="Custom Report Processor", layout="wide")

# ----------------------------
# LOGIN SYSTEM
# ----------------------------
# Set your credentials here
USERNAME = "dolphin"
PASSWORD = "Outsourcinghubindia@2025"

# Initialize session state for login
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

def login():
    st.title("🔒 Secure Login Required")
    st.write("Please enter your credentials to access the Custom Report Processor.")
    
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")
        
        if submitted:
            if username == USERNAME and password == PASSWORD:
                st.session_state.logged_in = True
                st.success("✅ Login successful! Redirecting...")
                st.rerun()
            else:
                st.error("❌ Invalid username or password. Please try again.")

# Show login page if not logged in
if not st.session_state.logged_in:
    login()
    st.stop()


st.set_page_config(page_title="Custom Report Processor", layout="wide")
st.title("📑 Custom Report Processor")

uploaded_file = st.file_uploader("Upload your Excel file (Custom Report.xlsx)", type=["xlsx"])




if uploaded_file is not None:
    try:
        file_bytes = uploaded_file.read()

        # Two workbook objects:
        # - wb: writable workbook (used for modifications)
        # - wb_data: read-only-like workbook with data_only=True to get computed values
        wb = load_workbook(io.BytesIO(file_bytes))
        wb_data = load_workbook(io.BytesIO(file_bytes), data_only=True)

        # load excel sheet (writeable)
        AR_Aging_sheet = wb["AR Aging (excluding HUD)"]
        Sample_sheet = wb["Sample Report"]
        Rent_sheet = wb["Rent Roll w. Lease Charges"]
        Legal_sheet = wb["Legal Report"]
        Tenant_memo_sheet = wb["Tenant Memo's"]

        # load excel sheet (data-only) for places where original code used data_only=True
        AR_Aging_sheet_data = wb_data["AR Aging (excluding HUD)"]
        Rent_sheet_data = wb_data["Rent Roll w. Lease Charges"]
        Legal_sheet_data = wb_data["Legal Report"]
        Tenant_memo_sheet_data = wb_data["Tenant Memo's"]
        Sample_sheet_data = wb_data["Sample Report"]

        # SAMPLE sheet header cells / fonts (kept same logic)
        ws = Sample_sheet
        # original had ws.font assignments; keep but guard in case Worksheet doesn't accept it
        try:
            ws.font = Font(bold=True)
            ws.font = Font(size=8)
        except Exception:
            # ignore if Worksheet doesn't support direct font assignment
            pass

        ws['J2'] = ("Unit")
        ws['H2'] = ("Name")
        ws['S2'] = ("Unit")

        # ----------------------------
        # THIS REPRESENTS UNITS IN THE LEGAL REPORT SHEET
        # ----------------------------
        sheet = Legal_sheet

        unit_legal = []
        blank_count = 0

        # Iterate column B starting from row 7
        for row in sheet.iter_rows(min_row=7, min_col=2, max_col=2):  # Column B from row 7
            cell = row[0]
            value = cell.value

            # Check if blank
            if value is None or str(value).strip() == "":
                blank_count += 1
                if blank_count > 1:  # Stop if more than 1 blank
                    break
                continue

            # Reset blank counter if non-blank value found
            blank_count = 0

            # Check if NOT bold
            if cell.font and not cell.font.bold:
                unit_legal.append(value)



        ws = Sample_sheet
        for col in range(0, len(unit_legal)):
            char = chr(65)  # 'A'
            ws[f"{char}{col + 3}"] = AR_Aging_sheet["A7"].value


        # ----------------------------
        # THIS REPRESENTS THE LIST HAVING ALL THE VALUES AVAILABLE IN THE UNIT COLUMN IN THE RENT SHEET
        # ----------------------------
        unit_rent = []

        def get_unit_rent_list(wb_obj, sheet_name):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"A{row}"].value

                if cell_value is None:
                    continue  # Skip empty cells

                cell_str = str(cell_value).strip().lower()

                if "summary" in cell_str:
                    break  # Stop loop when 'summary' is found

                unit_rent.append(cell_value)

            return unit_rent

        unit_rent = get_unit_rent_list(wb, "Rent Roll w. Lease Charges")

        # ----------------------------
        # PASTE THE DATA IN UNIT COLUMN OF LEGAL REPORT SHEET IN SAMPLE REPORT SHEET (column S)
        # (this corresponds to the earlier snippet that used chr(83) -> 'S')
        # ----------------------------
        ws = Sample_sheet
        for col in range(0, len(unit_legal)):
            char = chr(83)  # 'S'
            ws[f"{char}{col + 3}"] = unit_legal[col]

        # ----------------------------
        # AR AGING lists (A-F and related)
        # ----------------------------
        unit_ar = []
        unit_ar_cell = []

        # process column A (special logic in your original code)
        def process_until_total_A(wb_obj, sheet_name, column='A'):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"{column}{row}"].value
                cell_number = ws_local[f"{column}{row}"]
                if cell_value is not None and str(cell_value).strip().lower() == 'total':
                    break
                else:
                unit_ar.append(cell_value)
                unit_ar_cell.append(cell_number.coordinate)

        process_until_total_A(wb, "AR Aging (excluding HUD)", column='A')

        resident_ar = []

        def process_until_none_B(wb_obj, sheet_name, column='B'):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"{column}{row}"].value
                if cell_value is None:
                    break
                else:
                resident_ar.append(cell_value)

        process_until_none_B(wb, "AR Aging (excluding HUD)", column='B')

        status_ar = []

        def process_until_none_C(wb_obj, sheet_name, column='C'):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"{column}{row}"].value
                if cell_value is None:
                    break
                else:
                status_ar.append(cell_value)

        process_until_none_C(wb, "AR Aging (excluding HUD)", column='C')

        tenant_name_ar = []

        def process_until_none_D(wb_obj, sheet_name, column='D'):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"{column}{row}"].value
                if cell_value is None:
                    break
                else:
                tenant_name_ar.append(cell_value)

        process_until_none_D(wb, "AR Aging (excluding HUD)", column='D')

        _0_30_ar = []

        def process_until_none_F(wb_obj, sheet_name, column='F'):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"{column}{row}"].value
                if cell_value is None:
                    break
                else:
                _0_30_ar.append(cell_value)

        process_until_none_F(wb, "AR Aging (excluding HUD)", column='F')

        total_charges_ar = []

        def process_until_none_E(wb_obj, sheet_name, column='E'):
            ws_local = wb_obj[sheet_name]
            for row in range(8, ws_local.max_row + 1):
                cell_value = ws_local[f"{column}{row}"].value
                if cell_value is None:
                    break
                else:
                total_charges_ar.append(cell_value)

        process_until_none_E(wb, "AR Aging (excluding HUD)", column='E')

        # ----------------------------
        # PASTE VALUES OF UNIT AR AGING ACCORDING TO THE UNIT_LEGAL VALUES
        # ----------------------------
        ws = Sample_sheet

        row_start = 3
        current_row = row_start

        # Build AR dictionary for quick lookup
        ar_data = {}
        for u, resident, status, tenant_name, total_charges, ar_0_30 in zip(
                unit_ar, resident_ar, status_ar, tenant_name_ar, total_charges_ar, _0_30_ar):
            # preserve original normalization as provided
            key = str(u).strip().upper() if u is not None else ""
            ar_data[key] = {
                "resident": resident,
                "status": status,
                "tenant_name": tenant_name,
                "total_charges": total_charges,
                "ar_0_30": ar_0_30
            }

        # Track which units have been written in J column
        seen_units = set()

        # Track which units have already had ar_0_30 and total_charges written
        ar_amount_written_units = set()

        for unit in unit_legal:
            key = str(unit).strip().upper()

            # --- Column J: write unit_rent once, blank for repeats ---
            if key not in seen_units:
                ws[f"J{current_row}"] = key
                seen_units.add(key)
            else:
                ws[f"J{current_row}"] = ""

            # --- Columns B-E: repeatable AR info ---
            if key in ar_data:
                ws[f"B{current_row}"] = key                        # unit
                ws[f"C{current_row}"] = ar_data[key]["resident"]   # resident
                ws[f"D{current_row}"] = ar_data[key]["status"]     # status
                ws[f"E{current_row}"] = ar_data[key]["tenant_name"]# tenant name

                # --- Columns F-G: write only once per unit ---
                if key not in ar_amount_written_units:
                    ws[f"F{current_row}"] = ar_data[key]["ar_0_30"]
                    ws[f"G{current_row}"] = ar_data[key]["total_charges"]
                    ar_amount_written_units.add(key)
                # else leave F-G blank for repeated units

            current_row += 1

        # ----------------------------
        # FETCH RENT AMOUNTS (from Rent Roll) - original logic preserved
        # ----------------------------
        unit_rent_amount = []

        def fetch_rent_amounts(wb_obj, sheet_name):
            ws_local = wb_obj[sheet_name]
            unit_rent_amount_local = []
            for row in range(2, ws_local.max_row + 1):  # assuming row 1 = header
                g_value = ws_local[f"G{row}"].value
                h_value = ws_local[f"H{row}"].value

                # Normalize G text
                g_text = str(g_value).strip().lower() if g_value is not None else ""

                # Stop condition: if "market" is found
                if "market" in g_text:
                    break

                # Condition 1: G contains "rent"
                if "rent" in g_text:
                    if h_value is not None and isinstance(h_value, (int, float)) and h_value >= 0:
                        unit_rent_amount_local.append(h_value)

                # Condition 2: G is blank
                elif g_text == "":
                    if h_value in [0, 0.0]:  # only 0 or 0.00
                        unit_rent_amount_local.append(h_value)

            return unit_rent_amount_local

        # Use data-only workbook for amounts (like original which used data_only=True)
        unit_rent_amount = fetch_rent_amounts(wb_data, "Rent Roll w. Lease Charges")

        # ----------------------------
        # FETCH NAME RENT (original logic used property_value = Sample_sheet['A3'])
        # We'll preserve exact comparison behavior (cell object vs value) per original code
        # ----------------------------
        def fetch_name_rent(wb_obj, sheet_name, property_value=Sample_sheet['A3']):
            ws_local = wb_obj[sheet_name]
            name_rent_local = []
            for row in range(7, ws_local.max_row + 1):  # assuming row 1 = header
                value = ws_local[f"E{row}"].value

                # Stop condition: if cell value == property_value
                if value == property_value:
                    break

                # Append only if not blank
                if value is not None:
                    name_rent_local.append(value)

            return name_rent_local

        # call with data-only workbook (preserve original use of data_only)
        property_value = Sample_sheet['A3']
        name_rent = fetch_name_rent(wb_data, "Rent Roll w. Lease Charges", property_value)

        # ----------------------------
        # Map unit_rent -> names and amounts (zip)
        # Then write to Sample sheet columns J (unit), H (name), I (rent amount)
        # ----------------------------
        ws = Sample_sheet
        row_start = 3
        current_row = row_start

        # Prepare mappings for quick lookup
        unit_to_name = dict(zip(unit_rent, name_rent))
        unit_to_amount = dict(zip(unit_rent, unit_rent_amount))

        # Track which units have already had rent_amount written
        amount_written_units = set()

        for unit in unit_legal:
            key = str(unit).strip().upper()

            # --- Column J: unit_rent (repeat for every occurrence) ---
            ws[f"J{current_row}"] = key

            # --- Column H: name_rent (repeat for every occurrence) ---
            # Note: unit_to_name keys are not normalized in original code; we preserve that behavior
            if key in unit_to_name:
                ws[f"H{current_row}"] = unit_to_name[key]

            # --- Column I: rent amount (write only once per unit) ---
            if key in unit_to_amount and key not in amount_written_units:
                ws[f"I{current_row}"] = unit_to_amount[key]
                amount_written_units.add(key)
            # else leave blank

            current_row += 1

        # ----------------------------
        # TENANT MEMO related fetches
        # ----------------------------
        tenant_units = []

        def fetch_tenant_units(wb_obj, sheet_name="Tenant Memo's"):
            ws_local = wb_obj[sheet_name]
            tenant_units_local = []
            blank_count_local = 0  # consecutive blank counter

            for row in range(6, ws_local.max_row + 1):  # assuming row 1 = header
                value = ws_local[f"B{row}"].value

                if value is None or str(value).strip() == "":
                    blank_count_local += 1
                    if blank_count_local > 6:  # stop if more than 6 consecutive blanks
                        break
                    # If 1 or 2 blanks, just skip and continue
                    continue
                else:
                    tenant_units_local.append(value)  # append only non-blank
                    blank_count_local = 0  # reset blank counter

            return tenant_units_local

        # use data-only workbook for reading tenant memo values
        tenant_units = fetch_tenant_units(wb_data, "Tenant Memo's")

        unit_tenant_cell = []

        unit_rent_norm = [str(u).strip().upper() for u in unit_rent]
        tenant_units_norm = [str(u).strip().upper() for u in tenant_units]

        # List to store indices
        for unit in tenant_units_norm:
            if unit in unit_rent_norm:
                index = unit_rent_norm.index(unit)  # 0-based index
                unit_tenant_cell.append(index)


        # ----------------------------
        # fetch_type_tenant (search 'type' in column A)
        # ----------------------------
        def fetch_type_tenant(wb_obj, sheet_name="Tenant Memo's", keyword="type"):
            ws_local = wb_obj[sheet_name]
            type_tenant_local = []
            blank_count_local = 0  # consecutive blank counter

            for row in range(2, ws_local.max_row + 1):  # assuming row 1 = header
                value = ws_local[f"A{row}"].value

                if value is None or str(value).strip() == "":
                    blank_count_local += 1
                    if blank_count_local > 5:  # stop if more than 5 consecutive blanks
                        break
                    continue  # skip blank cell
                else:
                    blank_count_local = 0  # reset blank counter
                    # Check if the keyword is in the cell (case-insensitive)
                    if keyword.lower() in str(value).strip().lower():
                        type_tenant_local.append(value)

            return type_tenant_local

        type_tenant = fetch_type_tenant(wb_data)

        # ----------------------------
        # fetch_date_tenant (search 'date' in column C)
        # ----------------------------
        def fetch_date_tenant(wb_obj, sheet_name="Tenant Memo's", keyword="date"):
            ws_local = wb_obj[sheet_name]
            date_tenant_local = []
            blank_count_local = 0  # consecutive blank counter

            for row in range(2, ws_local.max_row + 1):  # assuming row 1 = header
                value = ws_local[f"C{row}"].value

                if value is None or str(value).strip() == "":
                    blank_count_local += 1
                    if blank_count_local > 5:  # stop if more than 5 consecutive blanks
                        break
                    continue  # skip blank cell
                else:
                    blank_count_local = 0  # reset blank counter
                    # Check if the keyword is in the cell (case-insensitive)
                    if keyword.lower() in str(value).strip().lower():
                        date_tenant_local.append(value)

            return date_tenant_local

        date_tenant = fetch_date_tenant(wb_data)

        # ----------------------------
        # fetch_memo_tenant (column D non-bold)
        # ----------------------------
        def fetch_memo_tenant(wb_obj, sheet_name="Tenant Memo's"):
            ws_local = wb_obj[sheet_name]
            memo_tenant_local = []
            blank_count_local = 0  # consecutive blank counter

            for row in range(2, ws_local.max_row + 1):  # assuming row 1 = header
                cell = ws_local[f"D{row}"]
                value = cell.value

                # Check for blank
                if value is None or str(value).strip() == "":
                    blank_count_local += 1
                    if blank_count_local > 5:  # stop if more than 5 consecutive blanks
                        break
                    continue  # skip blank cell
                else:
                    blank_count_local = 0  # reset blank counter

                # Check if the cell is not bold
                # (preserve original logic)
                if not cell.font.bold:
                    memo_tenant_local.append(value)

            return memo_tenant_local

        memo_tenant = fetch_memo_tenant(wb_data)

        # ----------------------------
        # Re-parse Legal sheet for unit_legal and status_legal (B-E), preserve original logic
        # ----------------------------
        ws_sample = Sample_sheet
        sheet = Legal_sheet

        unit_legal = []
        status_legal = []
        blank_count = 0

        # Iterate column B starting from row 7 (B to E)
        for row in sheet.iter_rows(min_row=7, min_col=2, max_col=5):  # B to E
            unit_cell = row[0]  # Column B
            status_cell = row[3]  # Column E

            unit_value = unit_cell.value
            status_value = status_cell.value

            # Check if blank
            if unit_value is None or str(unit_value).strip() == "":
                blank_count += 1
                if blank_count >= 20:  # stop if 20 consecutive blanks
                    break
                continue

            # Reset blank counter if non-blank found
            blank_count = 0

            # Check if NOT bold
            if unit_cell.font and not unit_cell.font.bold:
                unit_legal.append(str(unit_value).strip())
                status_legal.append(status_value)


        # --- Paste status into Sample Report column K ---
        row_start = 3
        for i, status in enumerate(status_legal):
            ws_sample[f"K{row_start + i}"] = status

        # ----------------------------
        # Map tenant memo lists into Sample sheet (T,U,V)
        # ----------------------------
        ws = Sample_sheet

        # Track units for which Tenant Memo info has been written
        memo_written_units = set()

        # Assuming tenant_units, type_tenant, date_tenant, memo_tenant are lists
        # And all lists are aligned (i.e., same order)
        tenant_data = dict()
        for unit, ttype, tdate, tmemo in zip(tenant_units, type_tenant, date_tenant, memo_tenant):
            key = str(unit).strip().upper()
            tenant_data[key] = {
                "type": ttype,
                "date": tdate,
                "memo": tmemo
            }

        row_start = 3
        current_row = row_start

        for unit in unit_legal:
            key = str(unit).strip().upper()

            if key in tenant_data and key not in memo_written_units:
                ws[f"T{current_row}"] = tenant_data[key]["type"]
                ws[f"U{current_row}"] = tenant_data[key]["date"]
                ws[f"V{current_row}"] = tenant_data[key]["memo"]
                memo_written_units.add(key)
            # else → leave blank if repeated or unit not in tenant_data

            current_row += 1

        # ----------------------------
        # LEGAL details (book_number, legal_type, etc.) reading from Legal_sheet and writing to Sample
        # ----------------------------
        sheet = Legal_sheet
        ws_sample = Sample_sheet

        # Lists to store fetched values
        book_number = []
        legal_type = []
        legal_reason = []
        active_legal = []
        due_date = []
        current_alert = []
        legal_notes = []

        # Track consecutive blanks
        blank_count = 0
        max_consecutive_blank = 100

        # Iterate starting from row 7 and columns F->R
        for row in sheet.iter_rows(min_row=8, min_col=6, max_col=18):  # F to R (max_col covers last needed)
            # Map columns
            f_cell = row[0]   # F -> book_number
            g_cell = row[1]   # G -> legal_type
            h_cell = row[2]   # H -> legal_reason
            i_cell = row[3]   # I -> active_legal
            m_cell = row[7]   # M -> due_date
            n_cell = row[8]   # N -> current_alert
            r_cell = row[12]  # R -> legal_notes

            cells = [f_cell, g_cell, h_cell, i_cell, m_cell, n_cell, r_cell]

            # Check if all are blank
            if all(cell.value is None or str(cell.value).strip() == "" for cell in cells):
                blank_count += 1
                if blank_count > max_consecutive_blank:
                    break
            else:
                blank_count = 0  # reset blank counter

            # Only fetch if NOT bold, otherwise append blank
            book_number.append(f_cell.value if not f_cell.font.bold else "")
            legal_type.append(g_cell.value if not g_cell.font.bold else "")
            legal_reason.append(h_cell.value if not h_cell.font.bold else "")
            active_legal.append(i_cell.value if not i_cell.font.bold else "")
            due_date.append(m_cell.value if not m_cell.font.bold else "")
            current_alert.append(n_cell.value if not n_cell.font.bold else "")
            legal_notes.append(r_cell.value if not r_cell.font.bold else "")

        # --- Paste values in Sample Report ---
        row_start = 3
        for i in range(len(book_number)):
            ws_sample[f"L{row_start + i}"] = book_number[i]
            ws_sample[f"M{row_start + i}"] = legal_type[i]
            ws_sample[f"N{row_start + i}"] = legal_reason[i]
            ws_sample[f"O{row_start + i}"] = active_legal[i]
            ws_sample[f"P{row_start + i}"] = due_date[i]
            ws_sample[f"Q{row_start + i}"] = current_alert[i]
            ws_sample[f"R{row_start + i}"] = legal_notes[i]


# Auto-adjust column widths for Sample_sheet safely
        for col in Sample_sheet.iter_cols():
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # Get column letter

            for cell in col:
                try:
                    if cell.value is not None:
                # Only measure actual text length, ignore formulas/styles
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

    # Keep a maximum width cap so very long text won't stretch sheet
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width > 0:
                Sample_sheet.column_dimensions[col_letter].width = adjusted_width


        # ----------------------------
        # Save processed workbook back to BytesIO and offer for download
        # ----------------------------
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Workbook has been processed successfully ✅")
        st.download_button(
            label="📥 Download Processed Report",
            data=output,
            file_name="Custom Report Processed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Error while processing the workbook: {e}")
        import traceback
        st.text(traceback.format_exc())
else:
    st.info("Please upload your Custom Report.xlsx file to process.")

