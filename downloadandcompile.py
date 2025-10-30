import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
import time
import os
import glob
from datetime import datetime

st.set_page_config(page_title="Report Downloader", layout="wide")
st.title("🏠 Yardi Automated Report Downloader")

st.write("Add property names to the list, then click **Start** to download all reports for each property.")

# --- Initialize session state ---
if 'property_list' not in st.session_state:
    st.session_state.property_list = []

# --- Input for adding property names ---
property_name_input = st.text_input("Enter Property Name")

col1, col2 = st.columns([1,1])
with col1:
    if st.button("➕ Add Property"):
        if property_name_input.strip():
            st.session_state.property_list.append(property_name_input.strip())
            st.success(f"Added: {property_name_input.strip()}")
        else:
            st.error("Please enter a valid property name before adding.")

with col2:
    if st.button("🧹 Clear List"):
        st.session_state.property_list = []
        st.warning("Property list cleared.")

# --- Display the current list ---
if st.session_state.property_list:
    st.subheader("📋 Property List:")
    for i, name in enumerate(st.session_state.property_list, 1):
        st.write(f"{i}. {name}")
else:
    st.info("No properties added yet.")

# --- Date Input ---
date_input = st.date_input("Select the Report Date")

# --- Helper function to rename downloaded files ---
def rename_latest_file(download_path, new_name, wait_time=30):
    """
    Waits for a new file to appear in the download_path and renames it.
    """
    seconds = 0
    file_renamed = False
    while seconds < wait_time:
        list_of_files = glob.glob(os.path.join(download_path, '*'))
        if list_of_files:
            latest_file = max(list_of_files, key=os.path.getctime)
            if latest_file.endswith(".crdownload"):  # Edge incomplete download
                time.sleep(1)
                seconds += 1
                continue
            new_file_path = os.path.join(download_path, new_name)
            os.rename(latest_file, new_file_path)
            file_renamed = True
            break
        time.sleep(1)
        seconds += 1
    return file_renamed

# --- Start Button ---
if st.button("🚀 Start Download"):
    if not st.session_state.property_list or not date_input:
        st.error("Please add at least one property and select a date.")
    else:
        date_to_fetch = date_input.strftime('%m/%d/%Y')
        MM_YY_date = date_input.strftime("%m/%Y")

        # === Selenium Config ===
        edge_driver_path = r"C:\Users\Custom\Downloads\edgedriver_win64\msedgedriver.exe"
        download_path = r"C:\Users\Custom\Desktop\Custom Report\downloaded_files"

        edge_options = Options()
        edge_options.use_chromium = True
        edge_options.add_argument("--start-maximized")
        prefs = {
            "download.default_directory": download_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        edge_options.add_experimental_option("prefs", prefs)
        service = Service(executable_path=edge_driver_path)
        driver = webdriver.Edge(service=service, options=edge_options)
        wait = WebDriverWait(driver, 60)

        # === LOGIN Function ===
        def login():
            try:
                driver.get("https://www.yardiasp14.com/66553dolphin/pages/LoginAdvanced.aspx")
                st.write("🔐 Waiting for login...")
                WebDriverWait(driver, 300).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "menuTab"))
                )
                st.success("✅ Logged in successfully.")
            except Exception as e:
                st.error(f"❌ Login Error: {e}")
                driver.quit()
                st.stop()

        # === REPORT FUNCTIONS (with renaming) ===
        def download_legal_report(property_name):
            try:
                st.write(f"📄 Downloading Legal Report for {property_name}...")

                driver.switch_to.default_content()

                # 👉 If property is affordable (i.e., NOT one of the standard properties), use alternate navigation
                if property_name.lower() not in ["bronxpa1", "bronxpa2", "bronxpa3"]:
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi0"]/a'))).click()
                    time.sleep(1)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm0"]/li[5]/a'))).click()
                    time.sleep(2)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi1"]/a'))).click()
                    time.sleep(1)
                else:
                    driver.find_element(By.ID, "mi1").click()

                # 🧭 Continue normal Legal Report logic
                sql_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "SQL Reports")))
                driver.execute_script("arguments[0].click();", sql_link)
                time.sleep(3)

                wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "filter")))

                search_box = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "k-input")))
                search_box.click()
                search_box.send_keys("Legal Report 2020")

                for _ in range(6):
                    driver.switch_to.active_element.send_keys(Keys.TAB)
                driver.switch_to.active_element.send_keys(Keys.RETURN)

                time.sleep(2)

                search_box_property = driver.find_element(By.ID, "hProp")
                search_box_property.clear()
                search_box_property.send_keys(property_name)
                search_box_property.send_keys(Keys.RETURN)
                search_box_property.send_keys(Keys.TAB)

                select_element = driver.find_element(By.NAME, 'tiStatus')
                select = Select(select_element)
                select.select_by_visible_text("Current")
                select_element.send_keys(Keys.TAB)

                driver.switch_to.active_element.send_keys(date_to_fetch)

                submit_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "filter_submit")))
                driver.execute_script("arguments[0].scrollIntoView(true);", submit_button)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", submit_button)

                excel_button = wait.until(EC.element_to_be_clickable((By.ID, "cmdExcel_Button")))
                excel_button.click()

                time.sleep(5)
                driver.switch_to.default_content()

                # 📝 Rename file safely
                filename = f"legal_report_{property_name}.xlsx"
                rename_latest_file(download_path, filename)
                st.write(f"✅ Legal Report downloaded and renamed: {filename}")

            except Exception as e:
                st.write(f"❌ Legal Report Error for {property_name}: {e}")


        def download_rent_roll(property_name):
            if property_name.lower() in ["bronxpa1", "bronxpa2", "bronxpa3"]:
                # === Standard Rent Roll (unchanged) ===
                try:
                    st.write(f"📄 Downloading Standard Rent Roll for {property_name}...")
                    actions = ActionChains(driver)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi1"]/a'))).click()
                    actions.move_to_element(wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi1-2"]/a')))).perform()
                    time.sleep(1)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm1-2"]/li[2]/a'))).click()
                    time.sleep(2)
                    driver.switch_to.frame(driver.find_elements(By.TAG_NAME, "iframe")[-1])
                    driver.find_element(By.ID, "PropLookup_LookupCode").clear()
                    driver.find_element(By.ID, "PropLookup_LookupCode").send_keys(property_name)
                    driver.find_element(By.ID, "Date2_TextBox").clear()
                    driver.find_element(By.ID, "Date2_TextBox").send_keys(date_to_fetch)
                    driver.find_element(By.ID, "MMYY2_TextBox").clear()
                    driver.find_element(By.ID, "MMYY2_TextBox").send_keys(MM_YY_date)
                    Select(driver.find_element(By.ID, "ReportType_DropDownList")).select_by_visible_text("Rent Roll with Lease Charges")
                    driver.find_element(By.ID, "Excel_Button").click()
                    time.sleep(5)
                    driver.switch_to.default_content()
                    filename = f"rent_roll_{property_name}.xlsx"
                    rename_latest_file(download_path, filename)
                    st.write(f"✅ Rent Roll downloaded and renamed: {filename}")
                except Exception as e:
                    st.write(f"❌ Rent Roll Error for {property_name}: {e}")
            else:
                # === Affordable Rent Roll with popup handling ===
                try:
                    st.write(f"📄 Downloading Affordable Rent Roll for {property_name}...")
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi0"]/a'))).click()
                    actions = ActionChains(driver)
                    actions.move_to_element(wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi0"]')))).perform()
                    time.sleep(0.8)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm0"]/li[5]/a'))).click()
                    time.sleep(2)

                    driver.switch_to.default_content()
                    driver.switch_to.frame(wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))[0])
                    time.sleep(0.8)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_ctl310"]'))).click()
                    time.sleep(1.5)
                    driver.switch_to.default_content()
                    driver.switch_to.frame(driver.find_elements(By.TAG_NAME, "iframe")[-1])

                    Select(driver.find_element(By.ID, "YsiMergeReport_DropDownList")).select_by_visible_text(
                        "Affordable Rent Roll with Lease Charges (AffRntRollLsChgs)"
                    )
                    time.sleep(4)
                    driver.switch_to.default_content()
                    driver.switch_to.frame(driver.find_elements(By.TAG_NAME, "iframe")[-1])

                    driver.find_element(By.ID, "Ysi4114_LookupCode").clear()
                    driver.find_element(By.ID, "Ysi4114_LookupCode").send_keys(property_name)
                    driver.find_element(By.ID, "Ysi4117_TextBox").clear()
                    driver.find_element(By.ID, "Ysi4117_TextBox").send_keys(date_to_fetch)
                    driver.find_element(By.ID, "Ysi4118_TextBox").clear()
                    driver.find_element(By.ID, "Ysi4118_TextBox").send_keys(MM_YY_date)
                    Select(driver.find_element(By.ID, "Ysi4122_DropDownList")).select_by_visible_text("Unit")
                    Select(driver.find_element(By.ID, "YsiOutpuType_DropDownList")).select_by_visible_text("Excel")

                    # Store main window handle before submit
                    main_window = driver.current_window_handle
                    handles_before = driver.window_handles

                    # Click Submit
                    wait.until(EC.element_to_be_clickable((By.ID, "btnSubmit_Button"))).click()

                    # Wait and handle popup
                    try:
                        WebDriverWait(driver, 10).until(EC.new_window_is_opened(handles_before))
                        popup_handle = [h for h in driver.window_handles if h not in handles_before]
                        if popup_handle:
                            driver.switch_to.window(popup_handle[0])
                            time.sleep(2)  # Give it time before closing
                            driver.close()
                            driver.switch_to.window(main_window)
                    except:
                        driver.switch_to.window(main_window)

                    driver.switch_to.default_content()

                    # Rename downloaded file
                    filename = f"affordable_rent_roll_{property_name}.xlsx"
                    rename_latest_file(download_path, filename)
                    st.write(f"✅ Affordable Rent Roll downloaded and renamed: {filename}")
                except Exception as e:
                    st.write(f"❌ Affordable Rent Roll Error for {property_name}: {e}")



        def download_tenant_memo(property_name):
            try:
                st.write(f"📄 Downloading Tenant Memo Report for {property_name}...")
                driver.switch_to.default_content()
                driver.find_element(By.ID, "mi1").click()
                resident_menu = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Resident")))
                ActionChains(driver).move_to_element(resident_menu).perform()
                tenant_memo_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Tenant Memos")))
                driver.execute_script("arguments[0].click();", tenant_memo_link)
                time.sleep(3)
                driver.switch_to.frame("filter")
                prop_element = driver.find_element(By.ID, "phMy")
                prop_element.clear()
                prop_element.send_keys(property_name)
                prop_element.send_keys(Keys.TAB)
                for _ in range(6):
                    driver.switch_to.active_element.send_keys(Keys.TAB)
                    time.sleep(0.2)
                driver.switch_to.active_element.send_keys(date_to_fetch)
                driver.find_element(By.XPATH, "//input[@value='Submit']").click()
                export_button = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewer1__ctl5__ctl4__ctl0_ButtonLink")))
                export_button.click()
                dropdown_option = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Excel')]")))
                dropdown_option.click()
                time.sleep(10)
                driver.switch_to.default_content()
                # --- Rename file ---
                filename = f"tenant_memo_{property_name}.xlsx"
                rename_latest_file(download_path, filename)
                st.write(f"✅ Tenant Memo Report downloaded and renamed: {filename}")
            except Exception as e:
                st.write(f"❌ Tenant Memo Error for {property_name}: {e}")


        def download_AR_AGING(property_name):
            try:
                st.write(f"📄 Downloading A/R Aging Report for {property_name}...")
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi0"]/a'))).click()
                actions = ActionChains(driver)
                actions.move_to_element(wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi0"]')))).perform()
                time.sleep(0.8)
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm0"]/li[5]/a'))).click()
                time.sleep(2)
                driver.switch_to.default_content()
                iframes = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
                driver.switch_to.frame(iframes[0])
                time.sleep(0.8)
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_ctl313"]'))).click()
                time.sleep(1.5)
                driver.switch_to.default_content()
                iframes = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
                driver.switch_to.frame(iframes[-1])
                driver.find_element(By.ID, "PropLookup_LookupCode").clear()
                driver.find_element(By.ID, "PropLookup_LookupCode").send_keys(property_name)
                tenant_status = driver.find_element(By.ID, "TenantStatus_MultiSelect")
                select = Select(tenant_status)
                select.select_by_visible_text("Current")
                driver.find_element(By.ID, "MMYY2_TextBox").clear()
                driver.find_element(By.ID, "MMYY2_TextBox").send_keys(MM_YY_date)
                aragingsummary = Select(driver.find_element(By.ID, "ReportType_DropDownList"))
                aragingsummary.select_by_visible_text("Receivable Aging Summary")
                driver.find_element(By.ID, "Excel_Button").click()
                time.sleep(7)
                driver.switch_to.default_content()
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi0"]/a'))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm0"]/li[5]/a'))).click()
                # --- Rename file ---
                filename = f"ar_aging_{property_name}.xlsx"
                rename_latest_file(download_path, filename)
                st.write(f"✅ A/R Aging Report downloaded and renamed: {filename}")
            except Exception as e:
                st.write(f"❌ A/R Aging Error for {property_name}: {e}")

        # --- EXECUTE LOGIN ONCE ---
        login()

        # --- EXECUTE REPORTS FOR EVERY PROPERTY ---
        for prop in st.session_state.property_list:
            download_rent_roll(prop)
            download_legal_report(prop)
            download_tenant_memo(prop)
            download_AR_AGING(prop)
        # =========================================================
        #              CONSOLIDATION SCRIPT (COMPLETE)
        # =========================================================
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import copy
        import pytz
        import os
        
        def consolidate_reports(property_name):
            timezone = pytz.timezone("US/Eastern")
            today_date = datetime.now(timezone).strftime("%m_%d_%Y")
        
            wb = Workbook()
            wb.remove(wb.active)
        
            # === Sheet names ===
            # Detect rent roll type
            is_standard = property_name.lower() in ["bronxpa1", "bronxpa2", "bronxpa3"]
            rent_roll_sheet_name = "Rent Roll w. Lease Charges" if is_standard else "Affordable Rent Roll"

            # === Sheet names ===
            sheet_names = [
                "Sample Report",
                "Support-->",
                "AR Aging (excluding HUD)",
                "Legal Report",
                rent_roll_sheet_name,
                "Tenant Memo's"
            ]

        
            for sheet in sheet_names:
                wb.create_sheet(title=sheet)
        
            ws = wb["Sample Report"]
        
            # === Shared Styles ===
            thin = Side(border_style="thin", color="000000")
            border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
            header_font = Font(name="Tahoma", size=8, bold=True)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
            # === Headers ===
            row1_titles = [
            "AR Aging (Excluding HUD)", None, None, None, None, None, None, None,
            rent_roll_sheet_name, None,
            "Legal Report", None, None, None, None, None, None, None, None,
            "Tenant Memo's", None, None
          ]

        
            row2_headers = [
                "Property", "Unit", "Tenant Code", "Status", "Tenant Name",
                "0-30 Day Balance(current due)", "Total Tenant Balance", "Name",
                "Current Rent", "Unit", "Status", "Book Number", "Legal Type", "Legal Reason",
                "Active Legal", "Due Date", "Current Alert", "Legal Notes", "Unit",
                "Memo Type", "Memo Date", "Memo"
            ]
        
            # === Fill headers ===
            num_cols = max(len(row1_titles), len(row2_headers))
            if len(row1_titles) < num_cols:
                row1_titles += [None] * (num_cols - len(row1_titles))
            if len(row2_headers) < num_cols:
                row2_headers += [""] * (num_cols - len(row2_headers))
        
            section_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
            for col_idx, value in enumerate(row1_titles, start=1):
                cell = ws.cell(row=1, column=col_idx, value=value)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border_all
                cell.fill = section_fill if value else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
            col = 1
            while col <= num_cols:
                title = row1_titles[col - 1]
                if title is None:
                    col += 1
                    continue
                start_col = col
                end_col = col
                while end_col + 1 <= num_cols and row1_titles[end_col] is None:
                    end_col += 1
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                col = end_col + 1
        
            for col_idx, value in enumerate(row2_headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border_all
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 18
            for col_idx in range(1, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            

            # === Input file paths ===
            download_dir = r"C:\Users\Custom\Desktop\Custom Report\downloaded_files"
            # Choose rent roll file based on property type
            if property_name.lower() in ["bronxpa1", "bronxpa2", "bronxpa3"]:
                rent_roll_file = os.path.join(download_dir, f"rent_roll_{property_name}.xlsx")
            else:
                rent_roll_file = os.path.join(download_dir, f"affordable_rent_roll_{property_name}.xlsx")
           
            sheet_file_map = {
                "AR Aging (excluding HUD)": os.path.join(download_dir, f"ar_aging_{property_name}.xlsx"),
                "Legal Report": os.path.join(download_dir, f"legal_report_{property_name}.xlsx"),
                rent_roll_sheet_name: rent_roll_file,
                "Tenant Memo's": os.path.join(download_dir, f"tenant_memo_{property_name}.xlsx")
            }


            # === Copy Data + Styles ===

            def copy_sheet_data_with_style(source_wb_path, target_ws):
                try:
                    src_wb = load_workbook(source_wb_path)
                    
                    # Pick correct source sheet: use first with data
                    sheet_candidates = [ws for ws in src_wb.worksheets if ws.max_row > 0 and ws.max_column > 0]
                    src_ws = sheet_candidates[0] if sheet_candidates else src_wb.active

                    # === Copy cell values and styles ===
                    for row in src_ws.iter_rows():
                        for cell in row:
                            tgt_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                tgt_cell.font = copy.copy(cell.font)
                                tgt_cell.fill = copy.copy(cell.fill)
                                tgt_cell.border = copy.copy(cell.border)
                                tgt_cell.alignment = copy.copy(cell.alignment)
                                tgt_cell.number_format = cell.number_format
                                tgt_cell.protection = copy.copy(cell.protection)

                    # === Copy merged cells ===
                    for merged_range in src_ws.merged_cells.ranges:
                        target_ws.merge_cells(str(merged_range))

                    # === Copy row dimensions ===
                    for idx, row_dim in src_ws.row_dimensions.items():
                        if row_dim.height is not None:
                            target_ws.row_dimensions[idx].height = row_dim.height
                        target_ws.row_dimensions[idx].hidden = row_dim.hidden

                    # === Copy column dimensions and hidden status ===
                    for col_letter, col_dim in src_ws.column_dimensions.items():
                        tgt_col_dim = target_ws.column_dimensions[col_letter]
                        if col_dim.width:
                            tgt_col_dim.width = col_dim.width
                        tgt_col_dim.hidden = col_dim.hidden  # 👈 THIS is the key fix

                    src_wb.close()
                    print(f"✅ Copied data and formatting from {source_wb_path}")

                except FileNotFoundError:
                    print(f"⚠️ File not found: {source_wb_path} — skipped.")
                except Exception as e:
                    print(f"❌ Error copying {source_wb_path}: {e}")


            # === Copy each report into target sheet ===
            for sheet_name, file_name in sheet_file_map.items():
                copy_sheet_data_with_style(file_name, wb[sheet_name])

            # === Save final workbook ===
            save_dir = r"C:\Users\Custom\Desktop\Custom Report"
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
            output_filename = os.path.join(save_dir, f"Custom_Report_{property_name}.xlsx")
            wb.save(output_filename)
            print(f"📁 '{output_filename}' created successfully.\n")
        
        # === Run this after each property's downloads ===
        for prop in st.session_state.property_list:
            consolidate_reports(prop)
        st.success("✅ All properties consolidated successfully!")

        driver.quit()
        st.success("🎉 All reports downloaded and renamed for all properties!")
